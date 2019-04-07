from openpyxl import load_workbook, Workbook


class Divider:
	def __init__(self):
		self.rb = Workbook()
		
	def get_xlsx(self, xname):
		return 
	
	def join(self, xname):
		print("从"+xname+"中读取数据")
		wb = load_workbook("%s.xlsx" % xname)
		sheets = wb.get_sheet_names()
		for sheet in sheets:
			for row in wb[sheet]["A"]:
				self.rb.active.append([row.value, ])
		return self.rb
	
	def divide(self, rb, d_num, f_name):
		sheet = rb.active
		sheet.title = "总表"
		a = list(sheet["A"])
		end_num = sheet.max_row % d_num
		print("判断行数是否是"+str(d_num)+"的整数")
		if end_num == 0:
			e = 1
		else:
			e = 2
		print("创建表并将数据保存")
		for xname in range(1, int(sheet.max_row / d_num) + e):
			print("判断剩余的数据量是否够取"+str(d_num)+"个")
			if len(a) > end_num:
				print("提取数据")
				group = a[0:d_num]
				print("创建新表"+str(xname))
				rb.create_sheet("分表" + str(xname))
				f_s = rb["分表" + str(xname)]
				print("删除总表中提取出来的数据")
				for i in group:
					a.remove(i)
				print("将数据添加到分表中")
				c_n = 1
				for i in group:
					f_s["A" + str(c_n)] = i.value
					c_n += 1
			else:
				if len(a) != 0:
					print("创建新表"+str(xname))
					rb.create_sheet("分表" + str(xname))
					f_s = rb["分表" + str(xname)]
					print("将数据添加到分表中")
					c_n = 1
					for i in a:
						f_s["A" + str(c_n)] = i.value
						c_n += 1
				else:
					pass
				# print(a)
		rb.save(f_name+".xlsx")


if __name__ == "__main__":
	d = Divider()
	wb = d.join("数据6_分表2")
	d.divide(wb, 2000, "数据6")
