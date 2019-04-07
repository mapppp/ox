from openpyxl import load_workbook


def divider(d, z, f):
	print("从"+z+"中读取数据")
	rb = load_workbook(z)
	sheet = rb.active
	sheet.title = "总表"
	a = list(sheet["A"])
	end_num = sheet.max_row % d
	print("判断行数是否是"+str(d)+"的整数")
	if end_num == 0:
		e = 1
	else:
		e = 2
	print("创建表并将数据保存")
	for xname in range(1, int(sheet.max_row / d) + e):
		print("判断剩余的数据量是否够取"+str(d)+"个")
		if len(a) > end_num:
			print("提取数据")
			group = a[0:d]
			print("创建新表"+str(xname))
			rb.create_sheet("分表" + str(xname))
			f_s = rb["分表" + str(xname)]
			print("删除总表中提取出来的数据")
			for i in group:
				a.remove(i)
			print("将数据添加到分表中")
			c_n = 1
			for i in group:
				f_s["A" + str(c_n)] = i.value
				c_n += 1
		else:
			if len(a) != 0:
				print("创建新表"+str(xname))
				rb.create_sheet("分表" + str(xname))
				f_s = rb["分表" + str(xname)]
				print("将数据添加到分表中")
				c_n = 1
				for i in a:
					f_s["A" + str(c_n)] = i.value
					c_n += 1
			else:
				pass
			# print(a)
	rb.save(f)


if __name__ == "__main__":
	d = 1000
	for z in ["数据6.18_19", "数据6.20_21", "数据6.22_23", "数据6.24_25"]:
		z = z + ".xlsx"
		f = z + "_分表.xlsx"
		divider(d, z, f)
