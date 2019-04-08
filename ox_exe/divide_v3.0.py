from openpyxl import load_workbook, Workbook
import os


class Divider:
    def __init__(self):
        self.xfile = {}

    def get_rb(self):
        list_dir = os.listdir(path='.')
        for directory in list_dir:
            if os.path.splitext(directory)[1] == ".xlsx":
                self.xfile[os.path.splitext(directory)[0]] = Workbook()
        return

    def join(self):
        for xname in self.xfile:
            print("从" + xname + "中读取数据")
            rb = load_workbook("%s.xlsx" % xname)
            sheets = rb.sheetnames
            for sheet in sheets:
                for row in rb[sheet]["A"]:
                        self.xfile[xname].active.append([row.value])
        return

    def divide(self, d_num):
        for xname in self.xfile:
            rb = self.xfile[xname]
            sheet = rb.active
            sheet.title = "总表"
            a = list(sheet["A"])
            end_num = sheet.max_row % d_num
            print("判断行数是否是"+str(d_num)+"的整数")
            if end_num == 0:
                e = 1
            else:
                e = 2
            print("创建表并将数据保存")
            for x in range(1, int(sheet.max_row / d_num) + e):
                print("判断剩余的数据量是否够取"+str(d_num)+"个")
                if len(a) > end_num:
                    print("提取数据")
                    group = a[0:d_num]
                    print("创建新表"+str(x))
                    rb.create_sheet("分表" + str(x))
                    f_s = rb["分表" + str(x)]
                    print("删除总表中提取出来的数据")
                    for i in group:
                        a.remove(i)
                    print("将数据添加到分表中")
                    c_n = 1
                    for i in group:
                        f_s["A" + str(c_n)] = i.value
                        c_n += 1
                else:
                    if len(a) != 0:
                        print("创建新表"+str(xname))
                        rb.create_sheet("分表" + str(xname))
                        f_s = rb["分表" + str(xname)]
                        print("将数据添加到分表中")
                        c_n = 1
                        for i in a:
                            f_s["A" + str(c_n)] = i.value
                            c_n += 1
                    else:
                        pass
                    # print(a)
            rb.save(xname+"_分表.xlsx")
        return


if __name__ == "__main__":
    num = int(input("输入每个分表的数据量："))
    d = Divider()
    d.get_rb()
    d.join()
    d.divide(num)
